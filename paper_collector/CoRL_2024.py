import requests
import pandas as pd
import time
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class CoRL2024PapersFetcher:
    """
    CoRL 2024论文抓取器，负责从OpenReview API获取论文数据。
    """
    def __init__(self, output_path, base_url="https://api2.openreview.net/notes", limit=25):
        self.output_path = output_path
        self.base_url = base_url
        self.headers = {
            "accept": "application/json",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        }
        self.params = {
            "content.venue": "CoRL 2024",
            "details": "replyCount,presentation",
            "domain": "robot-learning.org/CoRL/2024/Conference",
            "limit": limit,
            "offset": 0
        }
        self.papers = []

    def fetch_papers(self):
        """
        从OpenReview API抓取CoRL 2024论文数据。
        """
        while True:
            print(f"正在请求 offset={self.params['offset']} 的数据...")
            response = requests.get(self.base_url, headers=self.headers, params=self.params)
            if response.status_code != 200:
                print(f"请求失败，状态码：{response.status_code}")
                break

            data = response.json()
            notes = data.get("notes", [])

            if not notes:
                print("已获取所有数据！")
                break

            self._process_papers(notes)
            self.params["offset"] += 25  # 更新偏移量，获取下一页数据
            time.sleep(3)  # 为了避免请求过于频繁，暂停3秒

        self._save_papers_to_excel()

    def _process_papers(self, notes):
        """
        解析每一篇论文的信息，并添加到 papers 列表中。
        """
        for idx, note in enumerate(notes, start=len(self.papers) + 1):
            content = note.get("content", {})
            authors = content.get("authors", {}).get("value", [])
            abstract = content.get("abstract", {}).get("value", "")
            abstract = " ".join(abstract.split())  # 清除多余的空格和换行符，确保是一个连续的字符串
            self.papers.append({
                "Id": idx,
                "Title": content.get("title", {}).get("value", ""),
                "Authors": authors[0] if authors else "",
                "Abstract": abstract,
                "Video": content.get("video", {}).get("value", ""),
                "Website": content.get("website", {}).get("value", ""),
                "Code": content.get("code", {}).get("value", ""),
                "Original": f"https://openreview.net/attachment?id={note.get('forum', '')}&name=pdf"
            })

    def _save_papers_to_excel(self):
        """
        将抓取的论文数据保存为Excel文件。
        """
        df = pd.DataFrame(self.papers)
        df.to_excel(self.output_path, index=False)
        # print(f"数据已保存到 {self.output_path}")


class ExcelFormatter:
    """
    负责格式化Excel文件的类，主要用于设置样式、列宽、行高等。
    """
    def __init__(self, output_path):
        self.output_path = output_path
        self.wb = load_workbook(output_path)
        self.ws = self.wb.active

    def format_excel(self):
        """
        设置Excel文件的格式，包括列宽、行高、字体、对齐方式等。
        """
        self._set_column_widths()
        self._set_row_heights()
        self._freeze_header()
        self._set_font_styles()
        self._set_borders()
        self._save_excel()

    def _set_column_widths(self):
        """
        设置每列的宽度。
        """
        column_widths = [8.38, 35.88, 20, 82.38, 25, 25, 25, 25]
        for i, width in enumerate(column_widths, start=1):
            column_letter = self.ws.cell(row=1, column=i).column_letter
            self.ws.column_dimensions[column_letter].width = width

    def _set_row_heights(self):
        """
        设置每行的高度。
        """
        row_height = 200
        for row in self.ws.iter_rows():
            self.ws.row_dimensions[row[0].row].height = row_height

    def _freeze_header(self):
        """
        冻结标题行。
        """
        self.ws.freeze_panes = "A2"
        self.ws.row_dimensions[1].height = 26.25

    def _set_font_styles(self):
        """
        设置字体样式，包括标题行和其他单元格的字体。
        """
        font = Font(name="Times New Roman", size=14)
        header_font = Font(name="Times New Roman", size=14, bold=True)
        header_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")

        for cell in self.ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

        for row in self.ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        abstract_col_index = 4
        for row in self.ws.iter_rows(min_row=2, min_col=abstract_col_index, max_col=abstract_col_index):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _set_borders(self):
        """
        设置单元格边框样式。
        """
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = thin_border

    def _save_excel(self):
        """
        保存格式化后的Excel文件。
        """
        self.wb.save(self.output_path)
        print(f"已设置格式并保存到 {self.output_path}")


def main(output_path):
    fetcher = CoRL2024PapersFetcher(output_path)
    fetcher.fetch_papers()

    formatter = ExcelFormatter(output_path)
    formatter.format_excel()


if __name__ == "__main__":
    output_path = sys.argv[1]  # 从命令行参数获取输出文件路径
    main(output_path)
